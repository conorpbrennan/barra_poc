"""
barra_excel_check.py
====================
Excel validation workbook for a 3-position sub-book (top holdings at the latest COB).

Loads the cube's underlying frames and lays them out so that EVERY risk measure from
barra_factor_risk_cube.py is recomputed with live Excel formulas, next to pandas-computed
reference values for cross-checking:

  Net exposure x_k        = SUMPRODUCT(loadings_k, weights)
  Scenario PnL (per date) = SUMPRODUCT(x, factor_returns_row)          [dPnL = sum_k x_k*df_k]
  Scenario VaR 99         = -PERCENTILE.INC(PnL, 0.01)
  Scenario worst loss     = -MIN(PnL)
  Hypo shock PnL          = SUMPRODUCT(x, sigma_shock, factor_vols)    [vol via STDEV.S]
  Specific variance       = SUMPRODUCT(w^2, SpecificVar)               [diagonal block]
  Total VaR 99            = SQRT(VaR^2 + (2.326 * specific_vol)^2)

Conventions matching the cube: Market factor excluded from scenarios; a missing loading
contributes zero; SpecificVar is as-of joined (latest available per name).

    ../barra/bin/python barra_excel_check.py
"""
from __future__ import annotations
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from barra_factor_risk_cube import EVENT_WINDOWS, HYPO_SHOCKS, load_frames

OUT = pathlib.Path(__file__).resolve().parent.parent / "data"
XLSX = pathlib.Path(__file__).resolve().parent.parent / "tmp" / "barra_risk_check_3pos.xlsx"
N_POS = 3
Z99 = 2.326

PCT, NUM, SCI, DATE = "0.00%", "0.000", "0.00E+00", "yyyy-mm-dd"


def build() -> None:
    f = load_frames()
    exposures, positions, securities = f["exposures"], f["positions"], f["securities"]
    factor_ret, specific = f["factor_returns"], f["specific_var"]

    last = positions["Date"].max()
    book = (positions[positions["Date"] == last].nlargest(N_POS, "Weight")
            .merge(securities[["Position", "Ticker", "Issuer"]], on="Position"))
    figs = book["Position"].tolist()

    # wide factor-return history, ALL factors incl Market (leaf loading 1.0) — as build_scenarios()
    wide = (factor_ret
            .pivot(index="Date", columns="Factor", values="Return").dropna(how="any").sort_index())
    factors = list(wide.columns)

    # loadings at the as-of date; missing loading == 0 (matches the cube's empty leaf)
    L = (exposures[(exposures["Date"] == last) & (exposures["Position"].isin(figs))]
         .pivot(index="Position", columns="Factor", values="Loading")
         .reindex(index=figs, columns=factors).fillna(0.0))

    # SpecificVar as-of joined: latest available row per name on/before the as-of date
    sv = (specific[(specific["Position"].isin(figs)) & (specific["Date"] <= last)]
          .sort_values("Date").groupby("Position").last()
          .reindex(figs).rename(columns={"Date": "sv_asof", "SpecificVar": "sv"}))

    wts = book.set_index("Position")["Weight"].reindex(figs)

    # ---------------- pandas reference values (mirror the Excel formulas exactly) ----
    x = L.values.T @ wts.values                                   # net exposure per factor
    pnl = pd.Series(wide.values @ x, index=wide.index)            # dPnL = sum_k x_k*df_k
    vols = wide.std(ddof=1)                                       # == STDEV.S
    spec_var = float((wts.values ** 2 * sv["sv"].values).sum())
    spec_vol = float(np.sqrt(spec_var))

    def risk(p: pd.Series) -> tuple[float, float, float]:
        return -np.percentile(p, 1), -p.min(), p.mean()           # VaR99 / worst / mean

    ref = {"HistFull": risk(pnl)}
    for name, (a, b) in EVENT_WINDOWS.items():
        w = pnl.loc[a:b]
        if len(w):
            ref[name] = risk(w)
    for name, shock in HYPO_SHOCKS.items():
        h = float(sum(x[i] * shock.get(fc, 0.0) * vols[fc] for i, fc in enumerate(factors)))
        ref[name] = (-h, -h, h)

    # per-issuer standalone HistFull risk: each name held at its book weight w_i, so its
    # factor net-exposure is x_i,k = w_i * L_i,k and its PnL series is wide @ x_i.
    iss_ref = {}                                                  # fig -> (VaR, worst, mean, spec_vol, total)
    for fig in figs:
        xi = L.loc[fig].values * float(wts[fig])
        pnl_i = pd.Series(wide.values @ xi, index=wide.index)
        svol_i = float(np.sqrt(float(wts[fig]) ** 2 * float(sv.loc[fig, "sv"])))
        v, w_, m = risk(pnl_i)
        iss_ref[fig] = (v, w_, m, svol_i, float(np.sqrt(v ** 2 + (Z99 * svol_i) ** 2)))

    # ---- Level-2 contributions to HistFull Scenario VaR 99 (mirror the cube) ---------
    # Tail scenario t*: the cube uses quantile_index(book_pnl, 0.01, inc/lower) -> the
    # 0-based ascending index floor(0.01*(n-1)); contributions are evaluated on THAT day,
    # so Σ contributions = book P&L at t* (≈ Scenario VaR 99). var_q is the interpolated
    # quantile the cube divides by for "% of VaR".
    n_obs = len(pnl)
    var_rank = int(np.floor(0.01 * (n_obs - 1))) + 1              # 1-based rank for Excel SMALL()
    tail_pos = int(np.argsort(pnl.values, kind="stable")[var_rank - 1])
    df_tail = wide.iloc[tail_pos]                                 # factor returns on the tail day
    var_q = float(-np.percentile(pnl, 1))                         # = Scenario VaR 99 (interpolated)
    fac_comp = {fc: float(-x[i] * df_tail[fc]) for i, fc in enumerate(factors)}   # -x_k·df_k(t*)
    iss_comp = {fig: float(-float(wts[fig]) * (L.loc[fig].values @ df_tail.values)) for fig in figs}

    # ---------------- workbook --------------------------------------------------------
    wb = Workbook()
    bold = Font(bold=True)

    # --- Inputs ---
    ws = wb.active; ws.title = "Inputs"
    ws["A1"], ws["B1"] = "As-of date", last.date(); ws["B1"].number_format = DATE
    for c, h in enumerate(["Position", "Ticker", "Issuer", "Weight", "SpecificVar", "SV as-of"], 1):
        ws.cell(3, c, h).font = bold
    for r, fig in enumerate(figs, 4):
        rec = book.set_index("Position").loc[fig]
        ws.cell(r, 1, fig); ws.cell(r, 2, rec["Ticker"]); ws.cell(r, 3, rec["Issuer"])
        ws.cell(r, 4, float(wts[fig])).number_format = NUM
        ws.cell(r, 5, float(sv.loc[fig, "sv"])).number_format = SCI
        ws.cell(r, 6, sv.loc[fig, "sv_asof"].date()).number_format = DATE
    ws.cell(8, 1, "Weight sum (sub-book of the full Soros book)")
    ws.cell(8, 4, "=SUM(D4:D6)").number_format = NUM
    for col, w in zip("ABCDEF", (14, 8, 26, 10, 12, 12)):
        ws.column_dimensions[col].width = w

    # --- Exposures ---
    ws = wb.create_sheet("Exposures")
    ws.cell(1, 1, "Position").font = bold
    for c, fc in enumerate(factors, 2):
        ws.cell(1, c, fc).font = bold
    for r, fig in enumerate(figs, 2):
        ws.cell(r, 1, fig)
        for c, fc in enumerate(factors, 2):
            ws.cell(r, c, float(L.loc[fig, fc])).number_format = NUM
    ws.cell(6, 1, "Net exposure x_k (book)").font = bold
    for c in range(2, 2 + len(factors)):
        col = ws.cell(6, c).column_letter
        ws.cell(6, c, f"=SUMPRODUCT({col}2:{col}4,Inputs!$D$4:$D$6)").number_format = NUM
    # per-issuer net exposure x_i,k = loading_i,k * weight_i  (loading row 2+i, weight Inputs 4+i)
    tick = book.set_index("Position")["Ticker"]
    iss_x0 = 8
    for i, fig in enumerate(figs):
        ws.cell(iss_x0 + i, 1, f"x_k {tick[fig].upper()}")
        for c in range(2, 2 + len(factors)):
            col = ws.cell(iss_x0 + i, c).column_letter
            ws.cell(iss_x0 + i, c, f"={col}{2 + i}*Inputs!$D${4 + i}").number_format = NUM
    ws.column_dimensions["A"].width = 16

    # --- FactorReturns (+ live PnL columns) ---
    ws = wb.create_sheet("FactorReturns")
    nE = len(EVENT_WINDOWS)
    iss_pnl_c0 = len(factors) + 3 + nE                            # first per-issuer PnL column
    heads = (["Date"] + factors + ["PnL HistFull"]
             + [f"PnL {n.split(':')[1]}" for n in EVENT_WINDOWS]
             + [f"PnL {tick[fig].upper()}" for fig in figs])      # per-issuer PnL columns
    for c, h in enumerate(heads, 1):
        ws.cell(1, c, h).font = bold
    nfac = len(factors)
    pnl_col = ws.cell(1, nfac + 2).column_letter                  # "L" for 10 factors
    R = len(wide) + 1                                             # last data row
    for r, (dt, row) in enumerate(wide.iterrows(), 2):
        ws.cell(r, 1, dt.date()).number_format = DATE
        for c, fc in enumerate(factors, 2):
            ws.cell(r, c, float(row[fc])).number_format = PCT
        ws.cell(r, nfac + 2,
                f"=SUMPRODUCT(Exposures!$B$6:$K$6,B{r}:K{r})").number_format = PCT
        for j, (name, (a, b)) in enumerate(EVENT_WINDOWS.items()):
            a, b = pd.Timestamp(a), pd.Timestamp(b)
            ws.cell(r, nfac + 3 + j,
                    f'=IF(AND($A{r}>=DATE({a.year},{a.month},{a.day}),'
                    f'$A{r}<=DATE({b.year},{b.month},{b.day})),${pnl_col}{r},"")').number_format = PCT
        for i in range(len(figs)):                                # per-issuer PnL = x_i · returns_row
            irow = iss_x0 + i                                     # Exposures per-issuer x row
            ws.cell(r, iss_pnl_c0 + i,
                    f"=SUMPRODUCT(Exposures!$B${irow}:$K${irow},B{r}:K{r})").number_format = PCT
    ws.column_dimensions["A"].width = 11

    # --- Risk ---
    ws = wb.create_sheet("Risk")
    for c, fc in enumerate(factors, 2):
        ws.cell(1, c, fc).font = bold
    ws.cell(2, 1, "Factor vol σ (STDEV)").font = bold
    for c in range(2, 2 + nfac):
        col = ws.cell(2, c).column_letter
        # legacy STDEV (== STDEV.S, sample stdev): post-2007 names like STDEV.S are stored
        # by openpyxl without the required _xlfn. prefix -> #NAME? in Excel/LibreOffice.
        ws.cell(2, c, f"=STDEV(FactorReturns!{col}2:{col}{R})").number_format = PCT
    ws.cell(4, 1, "Hypo sigma shocks").font = bold
    hypo_rows = {}
    for i, (name, shock) in enumerate(HYPO_SHOCKS.items()):
        r = 5 + i; hypo_rows[name] = r
        ws.cell(r, 1, name)
        for c, fc in enumerate(factors, 2):
            ws.cell(r, c, float(shock.get(fc, 0.0)))
        ws.cell(r, nfac + 2,
                f"=SUMPRODUCT(Exposures!$B$6:$K$6,B{r}:K{r},$B$2:$K$2)").number_format = PCT
    ws.cell(4, nfac + 2, "Shock PnL").font = bold

    t0 = 10                                                       # scenario table header row
    for c, h in enumerate(["ScenarioSet", "VaR 99", "Worst loss", "Mean PnL", "Total VaR 99",
                           "VaR 99 (ref)", "Worst (ref)", "Mean (ref)"], 1):
        ws.cell(t0, c, h).font = bold
    scen_cols = {"HistFull": pnl_col}
    for j, name in enumerate(EVENT_WINDOWS):
        scen_cols[name] = ws.cell(1, nfac + 3 + j).column_letter
    r = t0 + 1
    sv_row, spec_rows = r + len(ref) + 2, None
    for name, (v_ref, w_ref, m_ref) in ref.items():
        ws.cell(r, 1, name)
        if name in scen_cols:                                      # vector scenarios: percentile/min/mean
            rng = f"FactorReturns!${scen_cols[name]}$2:${scen_cols[name]}${R}"
            ws.cell(r, 2, f"=-PERCENTILE({rng},0.01)")   # legacy == PERCENTILE.INC (inclusive)
            ws.cell(r, 3, f"=-MIN({rng})")
            ws.cell(r, 4, f"=AVERAGE({rng})")
        else:                                                      # hypo: length-1 vector == the shock PnL
            hr = hypo_rows[name]
            pl = ws.cell(hr, nfac + 2).coordinate
            ws.cell(r, 2, f"=-{pl}"); ws.cell(r, 3, f"=-{pl}"); ws.cell(r, 4, f"={pl}")
        ws.cell(r, 5, f"=SQRT(B{r}^2+({Z99}*$B${sv_row + 1})^2)")
        ws.cell(r, 6, v_ref); ws.cell(r, 7, w_ref); ws.cell(r, 8, m_ref)
        for c in range(2, 9):
            ws.cell(r, c).number_format = PCT
        r += 1

    ws.cell(sv_row, 1, "Specific variance Σw²σ²").font = bold
    ws.cell(sv_row, 2, "=SUMPRODUCT(Inputs!$D$4:$D$6*Inputs!$D$4:$D$6,Inputs!$E$4:$E$6)").number_format = SCI
    ws.cell(sv_row, 3, spec_var).number_format = SCI
    ws.cell(sv_row + 1, 1, "Specific vol").font = bold
    ws.cell(sv_row + 1, 2, f"=SQRT(B{sv_row})").number_format = PCT
    ws.cell(sv_row + 1, 3, spec_vol).number_format = PCT
    ws.cell(sv_row, 4, "(col C = pandas reference)")
    ws.column_dimensions["A"].width = 24

    # ---- per-issuer standalone risk (each name held at its book weight) -------------
    issuer = book.set_index("Position")["Issuer"]
    pi0 = sv_row + 3                                              # table header row
    ws.cell(pi0 - 1, 1, "Per-issuer standalone risk — HistFull (name at its book weight)").font = bold
    pihdr = ["Issuer", "VaR 99", "Worst loss", "Mean PnL", "Specific vol", "Total VaR 99",
             "VaR 99 (ref)", "Worst (ref)", "Mean (ref)", "Spec vol (ref)", "Total VaR (ref)"]
    for c, h in enumerate(pihdr, 1):
        ws.cell(pi0, c, h).font = bold
    for i, fig in enumerate(figs):
        rr = pi0 + 1 + i
        col = get_column_letter(iss_pnl_c0 + i)                  # FactorReturns per-issuer PnL col
        rng = f"FactorReturns!${col}$2:${col}${R}"
        wrow = 4 + i                                             # Inputs row for this name
        ws.cell(rr, 1, issuer[fig])
        ws.cell(rr, 2, f"=-PERCENTILE({rng},0.01)")
        ws.cell(rr, 3, f"=-MIN({rng})")
        ws.cell(rr, 4, f"=AVERAGE({rng})")
        ws.cell(rr, 5, f"=SQRT(Inputs!$D${wrow}^2*Inputs!$E${wrow})")   # w_i² · σ²_spec,i
        ws.cell(rr, 6, f"=SQRT(B{rr}^2+({Z99}*E{rr})^2)")              # Total VaR_i
        for c, val in enumerate(iss_ref[fig], 7):                # pandas refs: VaR/worst/mean/svol/total
            ws.cell(rr, c, val)
        for c in range(2, 12):
            ws.cell(rr, c).number_format = PCT

    # ---- Level-2 risk contributions to Scenario VaR 99 (HistFull) -------------------
    # Additive decomposition: evaluate each member's P&L on the BOOK's tail scenario t*
    # (the 1% VaR day), so the contributions sum to the book P&L at t* (= Scenario VaR 99).
    PCT1, NUM4 = "0.0%", "0.0000"
    ct0 = pi0 + len(figs) + 2
    ws.cell(ct0, 1, "Risk contributions to Scenario VaR 99 — HistFull "
                    "(additive; Σ = book VaR)").font = bold
    Lrng = f"FactorReturns!$L$2:$L${R}"
    ws.cell(ct0 + 1, 1, "VaR scenario rank (k-th worst)"); ws.cell(ct0 + 1, 2, var_rank)
    ws.cell(ct0 + 2, 1, "Book P&L at scenario t*")
    ws.cell(ct0 + 2, 2, f"=SMALL({Lrng},B{ct0 + 1})").number_format = PCT
    ws.cell(ct0 + 3, 1, "Scenario position (MATCH)")
    ws.cell(ct0 + 3, 2, f"=MATCH(B{ct0 + 2},{Lrng},0)")
    ws.cell(ct0 + 4, 1, "Scenario date")
    ws.cell(ct0 + 4, 2, f"=INDEX(FactorReturns!$A$2:$A${R},B{ct0 + 3})").number_format = DATE
    tp = f"$B${ct0 + 3}"                                          # 1-based position of t* in the data
    bvar = "$B$11"                                               # HistFull Scenario VaR 99 cell

    def contrib_table(t0, title, label_col, rows_spec):
        ws.cell(t0, 1, title).font = bold
        for c, hh in enumerate(["", "Net exp", "Component VaR", "Marginal VaR", "% of VaR",
                                "Comp (ref)", "Marg (ref)", "% (ref)"], 1):
            ws.cell(t0 + 1, c, hh).font = bold
        ws.cell(t0 + 1, 1, label_col).font = bold
        first = t0 + 2
        for k, (name, xexpr, dfexpr, cref, mref, pref) in enumerate(rows_spec):
            rr = first + k
            ws.cell(rr, 1, name)
            ws.cell(rr, 2, f"={xexpr}").number_format = NUM
            ws.cell(rr, 3, f"=-({xexpr})*({dfexpr})" if dfexpr else f"=-{xexpr}").number_format = PCT
            ws.cell(rr, 4, f"=C{rr}/B{rr}").number_format = NUM4
            ws.cell(rr, 5, f"=C{rr}/{bvar}").number_format = PCT1
            ws.cell(rr, 6, cref).number_format = PCT
            ws.cell(rr, 7, mref).number_format = NUM4
            ws.cell(rr, 8, pref).number_format = PCT1
        last = first + len(rows_spec) - 1
        ws.cell(last + 1, 1, "Σ  (= Scenario VaR 99)").font = bold
        ws.cell(last + 1, 3, f"=SUM(C{first}:C{last})").number_format = PCT
        ws.cell(last + 1, 5, f"=SUM(E{first}:E{last})").number_format = PCT1
        return last + 1

    # factor contributions: Component_k = -x_k · df_k(t*);  x_k = Exposures!<col>$6
    fac_spec = []
    for ci, fc in enumerate(factors):
        fcol = get_column_letter(2 + ci)                         # same factor col in Exposures & FactorReturns
        xexpr = f"Exposures!{fcol}$6"
        dfexpr = f"INDEX(FactorReturns!${fcol}$2:${fcol}${R},{tp})"
        cref = fac_comp[fc]
        mref = (fac_comp[fc] / x[ci]) if x[ci] else float("nan")
        fac_spec.append((fc, xexpr, dfexpr, cref, mref, fac_comp[fc] / var_q))
    fend = contrib_table(ct0 + 6, "Factor contributions", "Factor", fac_spec)

    # issuer contributions: Component_i = -PnL_i(t*);  Net exp x_i = Σ_k x_i,k (per-issuer row)
    it0 = fend + 2
    ws.cell(it0, 1, "Issuer contributions").font = bold
    for c, hh in enumerate(["Issuer", "Net exp", "Component VaR", "Marginal VaR", "% of VaR",
                            "Comp (ref)", "Marg (ref)", "% (ref)"], 1):
        ws.cell(it0 + 1, c, hh).font = bold
    for i, fig in enumerate(figs):
        rr = it0 + 2 + i
        pcol = get_column_letter(iss_pnl_c0 + i)
        ws.cell(rr, 1, issuer[fig])
        ws.cell(rr, 2, f"=SUM(Exposures!$B${8 + i}:$L${8 + i})").number_format = NUM
        ws.cell(rr, 3, f"=-INDEX(FactorReturns!${pcol}$2:${pcol}${R},{tp})").number_format = PCT
        ws.cell(rr, 4, f"=C{rr}/B{rr}").number_format = NUM4
        ws.cell(rr, 5, f"=C{rr}/{bvar}").number_format = PCT1
        xi_val = float(wts[fig]) * float(L.loc[fig].values.sum())
        ws.cell(rr, 6, iss_comp[fig]).number_format = PCT
        ws.cell(rr, 7, (iss_comp[fig] / xi_val) if xi_val else float("nan")).number_format = NUM4
        ws.cell(rr, 8, iss_comp[fig] / var_q).number_format = PCT1
    ilast = it0 + 1 + len(figs)
    ws.cell(ilast + 1, 1, "Σ  (= Scenario VaR 99)").font = bold
    ws.cell(ilast + 1, 3, f"=SUM(C{it0 + 2}:C{ilast})").number_format = PCT
    ws.cell(ilast + 1, 5, f"=SUM(E{it0 + 2}:E{ilast})").number_format = PCT1
    ws.cell(ct0, 4, "(cols F–H = pandas reference; t* = 1% VaR scenario)")

    wb.save(XLSX)
    print(f"wrote {XLSX}")
    print(f"book: {', '.join(f'{t.upper()} {w:.1%}' for t, w in zip(book['Ticker'], book['Weight']))}"
          f"  (as of {last.date()}, sub-book weight {wts.sum():.1%} of full book)")
    print(f"history: {len(wide)} periods x {nfac} factors | spec vol {spec_vol:.2%}\n")
    print(pd.DataFrame(ref, index=["VaR 99", "Worst loss", "Mean PnL"]).T
          .map(lambda v: f"{v:.2%}").to_string())


if __name__ == "__main__":
    build()
