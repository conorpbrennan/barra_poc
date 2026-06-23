"""
barra_excel_section3.py
=======================
Worked example: build ALL of **section 3** of the model reference (exposures z-scores ->
factor returns via cross-sectional WLS -> specific risk) for ONE name and ONE month, with the
whole estimation cross-section laid out and every step computed by LIVE EXCEL FORMULAS.

Section 3 is inherently cross-sectional — a z-score is taken against the universe's spread that
month, and the factor returns come from a regression across all names — so "one name, one month"
is a focus row sitting inside the full cross-section, which is all present so the formulas are real.

Sheets (mirroring barra_build_frames.py v2, §3 stages a–e):
  3a Raw          raw descriptors for the cross-section at month-end M (input values)
  3b-d Exposures  winsorised cross-sectional z-score per descriptor -> loadings; NonLinSize
                  orthogonalised to Size (c); the near-zero-spread factor screen (d) as a flag
  3e Regression   day-after-M stock returns regressed on the month-M loadings (weighted LS via
                  LINEST), giving factor returns (incl. Market = intercept) and the worked name's
                  residual u -> u², the specific-variance building block.

The transforms are exactly the v2 builder's:
  _winsor_z : med = MEDIAN; MAD = 1.4826·MEDIAN(|x−med|); z = clip((x−med)/MAD, ±3); re-standardise.
  regress   : daily WLS of stock returns on the prior month-end loadings, weight = exp(0.25·zSize),
              intercept column = the Market factor; residual² -> (EWMA over time) specific variance.

Reference columns (labelled "(ref)") carry the pandas value computed the same way, so the live
formula can be eyeballed against it. The z-score ties out to the production `exposures` frame for
the worked name; the regression ties out to `factor_returns` on the regression day. SpecificVar in
the frame is the EWMA of u² across the whole month's daily residuals — this sheet shows the single
day's u² that feeds that EWMA, not the smoothed value.

    ../barra/bin/python barra_excel_section3.py
"""
from __future__ import annotations
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

import barra_build_frames as bb
from barra_build_frames import (STYLE_FACTORS, MCAP_FLOOR, SOROS_CIK, START, END,
                                 _winsor_z, price_descriptors)
from barra_factor_risk_cube import load_frames

XLSX = pathlib.Path(__file__).resolve().parent.parent / "tmp" / "barra_section3_1name_1month.xlsx"
PCT, NUM, SCI, DATE = "0.00%", "0.0000", "0.00E+00", "yyyy-mm-dd"


# --------------------------------------------------------------------------- data pull (§3 inputs)
def pull_universe():
    """The estimation universe + raw price/fundamental pulls — the first half of
    barra_build_frames.build_frames(), reproduced so we can reach the RAW descriptors that the
    six output frames never store. Warm HTTP cache makes this cheap."""
    pos13f = bb.positions_from_13f(SOROS_CIK)
    xw = bb.crosswalk_cusips(pos13f["cusip"].tolist())
    t2c = bb.ticker_to_cik()
    sec = (xw.dropna(subset=["figi", "ticker"]).merge(t2c, on="ticker", how="left")
             .dropna(subset=["cik"]).drop_duplicates("figi"))
    extra_tickers = list(dict.fromkeys([t.lower() for t in bb.UNIVERSE_EXTRA] + bb.index_constituents()))
    extra = (pd.DataFrame({"ticker": extra_tickers}).merge(t2c, on="ticker", how="left")
               .dropna(subset=["cik"]))
    extra = extra[~extra["ticker"].isin(sec["ticker"])].copy()
    if not extra.empty:
        extra["figi"] = extra["ticker"]; extra["cusip"] = None
        sec = pd.concat([sec, extra], ignore_index=True).drop_duplicates("figi")
    sec = sec.head(bb.UNIVERSE_CAP).reset_index(drop=True)
    sec["cik"] = sec["cik"].astype(int)
    mkt = bb.stooq_daily(bb.MARKET_PROXY)
    prices = {t: p for t, p in bb._pull_map(bb.stooq_daily, sec["ticker"], "prices").items()
              if p is not None}
    funda = {int(c): f for c, f in
             bb._pull_map(bb.fundamentals, [int(c) for c in sec["cik"].unique()], "fundamentals").items()}
    return sec, prices, funda, mkt


def raw_descriptors(sec, prices, funda, mkt, cal):
    """The pre-z-score descriptor table (one row per name×date), exactly as build_exposures
    assembles it before the cross-sectional z-score. Copied from build_exposures so the raw values
    feeding the Excel z-score are identical to the model's."""
    pdsc = price_descriptors(prices, cal, mkt)
    frecs = []
    for _, row in sec.iterrows():
        f = funda.get(row["cik"]); px = prices.get(row["ticker"])
        if f is None or px is None:
            continue
        cl = px["Close"].reindex(cal, method="ffill")
        fa = pd.merge_asof(pd.DataFrame({"filed": cal}), f, on="filed", direction="backward")
        fa.index = cal
        mcap = (cl * fa["Shares"]).where(lambda m: m > MCAP_FLOOR)
        frecs.append(pd.DataFrame({
            "ticker": row["ticker"], "Date": cal,
            "Size": np.log(mcap + 1), "NonLinSize": np.log(mcap + 1) ** 3,
            "Value": fa["Equity"].values / (mcap.values + 1),
            "EarnYield": fa["NetIncome"].values / (mcap.values + 1),
            "Leverage": fa["Assets"].values / (fa["Equity"].values + 1),
            "Growth": (f["Assets"].pct_change().reindex(range(len(cal)), method=None).values
                       if "Assets" in f else np.nan),
            "mcap": mcap.values,
        }))
    fund = pd.concat(frecs, ignore_index=True) if frecs else pd.DataFrame()
    raw = pdsc.merge(fund, on=["ticker", "Date"], how="outer")
    raw = raw.merge(sec[["ticker", "figi", "title" if "title" in sec else "ticker"]]
                    .rename(columns={"title": "Issuer"}) if "title" in sec
                    else sec[["ticker", "figi"]], on="ticker", how="left")
    return raw.dropna(subset=["figi"])


# --------------------------------------------------------------------------- workbook helpers
def _winsor_z_block(ws, hdr_row, name_row0, n, src_col, stat_col, name):
    """Emit the live winsorised-z pipeline for ONE descriptor laid out in `src_col` (rows
    name_row0..name_row0+n-1). Writes med/MAD/std stats at the top (rows hdr_row..hdr_row+2 in
    `stat_col`) and the per-name clipped-z then re-standardised loading in the two columns right of
    src. Returns the loading column letter. Mirrors _winsor_z exactly (med/MAD/±3/re-standardise)."""
    r0, r1 = name_row0, name_row0 + n - 1
    src = get_column_letter(src_col)
    zc = get_column_letter(src_col)                      # placeholder; caller lays z/loading cols
    rng = f"{src}{r0}:{src}{r1}"
    med = f"{stat_col}{hdr_row}"; mad = f"{stat_col}{hdr_row + 1}"
    ws[med] = f"=MEDIAN({rng})"
    # MAD over non-blank cells -> sigma; array formula so blanks (missing descriptor) are excluded.
    ws[mad] = ArrayFormula(mad, f"=1.4826*MEDIAN(IF({rng}<>\"\",ABS({rng}-{med})))")
    for c in (med, mad):
        ws[c].number_format = NUM
    return med, mad, (r0, r1), rng


def build():
    frames = load_frames()
    exp_frame, fr_frame = frames["exposures"], frames["factor_returns"]
    sv_frame, sec_frame = frames["specific_var"], frames["securities"]

    sec, prices, funda, mkt = pull_universe()
    cal = pd.date_range(START, END, freq="ME")
    raw = raw_descriptors(sec, prices, funda, mkt, cal)

    # choose M = a month-end with a following regression day inside the sample (so §3e exists) and a
    # clean cross-section; use the second-to-last exposure month-end. d1 = first trading day after M.
    exp_dates = sorted(exp_frame["Date"].unique())
    M = pd.Timestamp(exp_dates[-2])
    R = pd.DataFrame({t: px["Close"].pct_change() for t, px in prices.items()}).loc[START:END]
    R = R.where(R.abs() <= 0.5)
    d1 = R.index[R.index > M][0]

    # ---- §3a/b reference: the model's loadings at M, recomputed here to tie out ------------------
    rawM = raw[raw["Date"] == M].copy()
    fig2tkr = sec.set_index("figi")["ticker"].to_dict()
    fig2iss = sec_frame.set_index("Position")["Issuer"].to_dict()
    # pandas mirror of build_exposures' per-date z-score (for the (ref) columns + tie-out)
    zM = rawM.set_index("figi")[STYLE_FACTORS].astype(float).copy()
    loadM = pd.DataFrame({f: _winsor_z(zM[f]) for f in STYLE_FACTORS})
    if {"NonLinSize", "Size"}.issubset(loadM):
        b = np.polyfit(loadM["Size"].fillna(0), loadM["NonLinSize"].fillna(0), 1)
        loadM["NonLinSize"] = _winsor_z(loadM["NonLinSize"] - (b[0] * loadM["Size"] + b[1]))

    # ---- §3e reference: reproduce the builder's WLS for (d0=M -> day d1) --------------------------
    Xd = loadM.copy()
    Xd = Xd[Xd.notna().sum(axis=1) >= 6].fillna(0.0)
    Xd = Xd.loc[:, Xd.std() > 0.05]                       # factor screen (d)
    screen_factors = list(Xd.columns)
    tk = pd.Series({fig: fig2tkr.get(fig) for fig in Xd.index})
    Xd = Xd[tk.reindex(Xd.index).isin(R.columns).values]
    figs = list(Xd.index)
    y = R.loc[d1, tk[figs].values].values
    ok = ~np.isnan(y)
    figs = list(np.array(figs)[ok]); Xd = Xd.iloc[ok]; y = y[ok]
    Xm = np.column_stack([np.ones(len(figs)), Xd.values])
    W = np.sqrt(np.exp(Xd["Size"].values) ** 0.5)
    beta, *_ = np.linalg.lstsq(Xm * W[:, None], y * W, rcond=None)
    fac_names = ["Market"] + list(Xd.columns)
    resid = y - Xm @ beta

    # worked name: the largest Soros weight present in the regression population at M
    pos = frames["positions"]
    posM = pos[pos["Date"] == M].set_index("Position")["Weight"]
    cand = [f for f in figs if f in posM.index]
    worked = (posM.loc[cand].sort_values(ascending=False).index[0] if cand else figs[0])
    wi = figs.index(worked)
    fr_ref = fr_frame[fr_frame["Date"] == d1].set_index("Factor")["Return"].to_dict()

    # ------------------------------------------------------------------ workbook
    wb = Workbook()
    bold = Font(bold=True); hl = PatternFill("solid", fgColor="FFF3B0")
    NAMEROW0 = 6                                          # first name row on the cross-section sheets

    # === 3a · Raw descriptors (input values) =====================================================
    ws = wb.active; ws.title = "3a Raw"
    ws["A1"] = "§3a · Raw descriptors — cross-section at month-end M (model inputs, pre-z-score)"
    ws["A1"].font = bold
    ws["A2"] = "Month-end M"; ws["B2"] = M.date(); ws["B2"].number_format = DATE
    ws["A3"] = "Worked name"; ws["B3"] = f"{fig2tkr.get(worked,'')}  ({fig2iss.get(worked,'')})"
    hdr = ["Position", "Ticker"] + STYLE_FACTORS + ["mcap"]
    for c, h in enumerate(hdr, 1):
        ws.cell(NAMEROW0 - 1, c, h).font = bold
    raw_figs = list(rawM["figi"])
    raw_by_fig = rawM.set_index("figi")
    row_of = {}
    for i, fig in enumerate(raw_figs):
        r = NAMEROW0 + i; row_of[fig] = r
        ws.cell(r, 1, fig); ws.cell(r, 2, fig2tkr.get(fig, ""))
        rec = raw_by_fig.loc[fig]
        for c, f in enumerate(STYLE_FACTORS, 3):
            v = rec[f]
            if pd.notna(v):
                ws.cell(r, c, float(v)).number_format = NUM
        mc = rec.get("mcap", np.nan)
        if pd.notna(mc):
            ws.cell(r, 3 + len(STYLE_FACTORS), float(mc)).number_format = SCI
        if fig == worked:
            for c in range(1, len(hdr) + 1):
                ws.cell(r, c).fill = hl
    nrows = len(raw_figs)
    last = NAMEROW0 + nrows - 1
    RAWCOL = {f: get_column_letter(3 + i) for i, f in enumerate(STYLE_FACTORS)}
    ws.column_dimensions["A"].width = 16; ws.column_dimensions["B"].width = 9

    # === 3b–d · Exposures: live winsorised z-score per descriptor ================================
    we = wb.create_sheet("3b-d Exposures")
    we["A1"] = "§3b–d · Cross-sectional winsorised z-score -> loadings (live formulas)"; we["A1"].font = bold
    we["A2"] = ("Per descriptor: med = MEDIAN; MAD = 1.4826·MEDIAN(|x−med|) (array, blanks excluded); "
                "z = clip((x−med)/MAD, ±3); loading = (z − mean)/std.  (d) screen flags std ≤ 0.05.")
    we["A3"] = "Stats row →"; we["A3"].font = bold
    for c in (("med", 3), ("MAD", 4), ("screen std", 5)):
        pass
    we.cell(3, 1, "med / MAD / screen below each factor")
    # header row of factor names at NAMEROW0-1; per factor: zclip col and loading col
    we.cell(NAMEROW0 - 1, 1, "Position").font = bold
    LOADCOL = {}
    col = 2
    for f in STYLE_FACTORS:
        zcol = get_column_letter(col); lcol = get_column_letter(col + 1)
        LOADCOL[f] = lcol
        we.cell(NAMEROW0 - 1, col, f"z {f}").font = bold
        we.cell(NAMEROW0 - 1, col + 1, f).font = bold
        rawrng = f"'3a Raw'!{RAWCOL[f]}{NAMEROW0}:{RAWCOL[f]}{last}"
        med = f"{zcol}3"; mad = f"{zcol}4"; scr = f"{zcol}5"
        we[med] = f"=MEDIAN({rawrng})"
        we[mad] = ArrayFormula(mad, f'=1.4826*MEDIAN(IF({rawrng}<>"",ABS({rawrng}-{med})))')
        we[med].number_format = NUM; we[mad].number_format = NUM
        # per-name clipped robust z (blank propagates), then loading = re-standardised clipped z
        for i, fig in enumerate(raw_figs):
            r = NAMEROW0 + i
            src = f"'3a Raw'!{RAWCOL[f]}{r}"
            we.cell(r, col).value = f'=IF({src}="","",MAX(-3,MIN(3,({src}-${med})/${mad})))'
            we.cell(r, col).number_format = NUM
            zrng = f"{zcol}{NAMEROW0}:{zcol}{last}"
            we.cell(r, col + 1).value = (
                f'=IF({zcol}{r}="","",({zcol}{r}-AVERAGE({zrng}))/(STDEV({zrng})+1E-12))')
            we.cell(r, col + 1).number_format = NUM
        lrng = f"{lcol}{NAMEROW0}:{lcol}{last}"
        we[scr] = f"=STDEV({lrng})"; we[scr].number_format = NUM
        we.cell(5, 1, "screen std (drop ≤0.05) →")
        col += 2
    for i, fig in enumerate(raw_figs):
        r = NAMEROW0 + i
        we.cell(r, 1, fig)
        if fig == worked:
            for c in range(1, col + 1):
                we.cell(r, c).fill = hl

    # NonLinSize orthogonalised to Size (c): residual of zNonLin on zSize, then re-winsor-z.
    # cols: slope, intercept, resid, rz (clipped robust z of resid), final loading.
    oc = col
    sl, ic2 = get_column_letter(oc), get_column_letter(oc + 1)
    rc, rzc, fc = get_column_letter(oc + 2), get_column_letter(oc + 3), get_column_letter(oc + 4)
    we.cell(NAMEROW0 - 1, oc, "NonLinSize ⟂Size: slope").font = bold
    we.cell(NAMEROW0 - 1, oc + 1, "intercept").font = bold
    we.cell(NAMEROW0 - 1, oc + 2, "resid").font = bold
    we.cell(NAMEROW0 - 1, oc + 3, "z resid").font = bold
    we.cell(NAMEROW0 - 1, oc + 4, "NonLinSize (final)").font = bold
    zN = f"{LOADCOL['NonLinSize']}{NAMEROW0}:{LOADCOL['NonLinSize']}{last}"
    zS = f"{LOADCOL['Size']}{NAMEROW0}:{LOADCOL['Size']}{last}"
    we[f"{sl}3"] = f"=SLOPE({zN},{zS})"; we[f"{ic2}3"] = f"=INTERCEPT({zN},{zS})"
    we[f"{sl}3"].number_format = NUM; we[f"{ic2}3"].number_format = NUM
    rrng = f"{rc}{NAMEROW0}:{rc}{last}"; rzrng = f"{rzc}{NAMEROW0}:{rzc}{last}"
    rmed, rmad = f"{rc}3", f"{rc}4"
    we[rmed] = f"=MEDIAN({rrng})"
    we[rmad] = ArrayFormula(rmad, f'=1.4826*MEDIAN(IF({rrng}<>"",ABS({rrng}-{rmed})))')
    we[rmed].number_format = NUM; we[rmad].number_format = NUM
    for i, fig in enumerate(raw_figs):
        r = NAMEROW0 + i
        ln = f"{LOADCOL['NonLinSize']}{r}"; ls = f"{LOADCOL['Size']}{r}"
        we.cell(r, oc + 2).value = f'=IF({ln}="","",{ln}-(${sl}$3*{ls}+${ic2}$3))'
        we.cell(r, oc + 2).number_format = NUM
        we.cell(r, oc + 3).value = f'=IF({rc}{r}="","",MAX(-3,MIN(3,({rc}{r}-${rmed})/${rmad})))'
        we.cell(r, oc + 3).number_format = NUM
        we.cell(r, oc + 4).value = (
            f'=IF({rzc}{r}="","",({rzc}{r}-AVERAGE({rzrng}))/(STDEV({rzrng})+1E-12))')
        we.cell(r, oc + 4).number_format = NUM
    LOADCOL["NonLinSize_final"] = fc
    we.column_dimensions["A"].width = 16

    # tie-out (ref) block: pandas loadings for the worked name vs what the formulas compute
    tcol = oc + 7
    we.cell(NAMEROW0 - 1, tcol, "Worked-name loading (ref, pandas)").font = bold
    we.cell(NAMEROW0 - 1, tcol + 1, "= frame loading?").font = bold
    fr_load = exp_frame[(exp_frame["Date"] == M) & (exp_frame["Position"] == worked)] \
        .set_index("Factor")["Loading"].to_dict()
    for i, f in enumerate(STYLE_FACTORS):
        r = NAMEROW0 + i
        we.cell(r, tcol - 1, f).font = bold
        ref = loadM.loc[worked, f] if worked in loadM.index else np.nan
        if pd.notna(ref):
            we.cell(r, tcol, float(ref)).number_format = NUM
        fv = fr_load.get(f)
        if fv is not None:
            we.cell(r, tcol + 1, float(fv)).number_format = NUM

    # === 3e · Weighted cross-sectional regression (LINEST) =======================================
    wr = wb.create_sheet("3e Regression")
    wr["A1"] = "§3e · Daily WLS of stock returns on the month-M loadings — factor returns + residual"
    wr["A1"].font = bold
    wr["A2"] = "Regression day d (first trading day after M)"; wr["B2"] = d1.date()
    wr["B2"].number_format = DATE
    wr["A3"] = (f"Population: {len(figs)} names with ≥6/10 loadings, factor screen std>0.05, return on d, "
                "Size present. Weight W = exp(0.25·zSize). Intercept column = Market factor.")
    wr["A4"] = ("LINEST(y·W, [W, loadings·W], FALSE) — const FALSE because the Market intercept is an "
                "explicit W column; coefficients come back in REVERSE column order (read right→left).")
    h0 = 6
    cols = ["Position", "Ticker", "y = ret(d)", "W"] + [f"L:{f}" for f in screen_factors] + \
           ["Market·W"] + [f"{f}·W" for f in screen_factors] + ["y·W", "fit ŷ", "resid u", "u²"]
    for c, h in enumerate(cols, 1):
        wr.cell(h0, c, h).font = bold
    base = h0 + 1
    # column letters
    cY = get_column_letter(3); cW = get_column_letter(4)
    cL = {f: get_column_letter(5 + i) for i, f in enumerate(screen_factors)}
    cMW = get_column_letter(5 + len(screen_factors))
    cXW = {f: get_column_letter(6 + len(screen_factors) + i) for i, f in enumerate(screen_factors)}
    cYW = get_column_letter(6 + 2 * len(screen_factors))
    cFIT = get_column_letter(7 + 2 * len(screen_factors))
    cU = get_column_letter(8 + 2 * len(screen_factors))
    cU2 = get_column_letter(9 + 2 * len(screen_factors))
    r1 = base + len(figs) - 1
    for i, fig in enumerate(figs):
        r = base + i
        wr.cell(r, 1, fig); wr.cell(r, 2, fig2tkr.get(fig, ""))
        wr.cell(r, 3, float(y[i])).number_format = PCT                 # input: realised return
        # W = exp(0.25 · zSize loading); reference the Exposures loading
        wr.cell(r, 4).value = f"=EXP(0.25*'3b-d Exposures'!{LOADCOL['Size']}{row_of[fig]})"
        wr.cell(r, 4).number_format = NUM
        for f in screen_factors:
            lc = LOADCOL["NonLinSize_final"] if f == "NonLinSize" else LOADCOL[f]
            src = f"'3b-d Exposures'!{lc}{row_of[fig]}"
            wr.cell(r, 5 + screen_factors.index(f)).value = f'=IF({src}="",0,{src})'   # zero-fill
            wr.cell(r, 5 + screen_factors.index(f)).number_format = NUM
        wr.cell(r, 5 + len(screen_factors)).value = f"=${cW}{r}"        # Market·W = 1·W
        wr.cell(r, 5 + len(screen_factors)).number_format = NUM
        for j, f in enumerate(screen_factors):
            wr.cell(r, 6 + len(screen_factors) + j).value = f"={cL[f]}{r}*${cW}{r}"
            wr.cell(r, 6 + len(screen_factors) + j).number_format = NUM
        wr.cell(r, 6 + 2 * len(screen_factors)).value = f"={cY}{r}*${cW}{r}"
        wr.cell(r, 6 + 2 * len(screen_factors)).number_format = NUM
        if fig == worked:
            for c in range(1, len(cols) + 1):
                wr.cell(r, c).fill = hl

    # LINEST over the weighted design [Market·W, L1·W, ... Lk·W]; const FALSE.
    xw_first, xw_last = cMW, cXW[screen_factors[-1]]
    linest_rng = f"{xw_first}{base}:{xw_last}{r1}"
    yw_rng = f"{cYW}{base}:{cYW}{r1}"
    k = 1 + len(screen_factors)
    lr = r1 + 3
    wr.cell(lr - 1, 1, "LINEST coefficients (reverse column order →)").font = bold
    coef_cells = [f"{get_column_letter(2 + j)}{lr}" for j in range(k)]
    wr.cell(lr, 1, "betas:")
    # one array formula spilling across k cells
    wr[coef_cells[0]] = ArrayFormula(f"{coef_cells[0]}:{coef_cells[-1]}",
                                     f"=LINEST({yw_rng},{linest_rng},FALSE)")
    for cc in coef_cells:
        wr[cc].number_format = NUM
    # map reversed LINEST output -> factor order; design cols are [Market, screen_factors...]
    design = ["Market"] + screen_factors
    fac_row = lr + 2
    wr.cell(fac_row - 1, 1, "Factor return (this day) — formula vs frame (ref)").font = bold
    wr.cell(fac_row, 1, "Factor").font = bold
    wr.cell(fac_row, 2, "from LINEST").font = bold
    wr.cell(fac_row, 3, "frame (ref)").font = bold
    for i, f in enumerate(design):
        rr = fac_row + 1 + i
        wr.cell(rr, 1, f)
        rev = coef_cells[k - 1 - i]                       # reverse mapping
        wr.cell(rr, 2, f"={rev}").number_format = PCT
        if f in fr_ref:
            wr.cell(rr, 3, float(fr_ref[f])).number_format = PCT
    coef_for = {f: coef_cells[k - 1 - i] for i, f in enumerate(design)}

    # residual + u² for every name. fit ŷ = Market_coef·1 + Σ_f beta_f·loading_f
    # (loadings are the unweighted cL columns; the regression solved on the W-scaled design).
    for i, fig in enumerate(figs):
        r = base + i
        parts = [f"${coef_for['Market']}"]
        parts += [f"${coef_for[f]}*{cL[f]}{r}" for f in screen_factors]
        wr.cell(r, 7 + 2 * len(screen_factors)).value = "=" + "+".join(parts)
        wr.cell(r, 7 + 2 * len(screen_factors)).number_format = PCT
        wr.cell(r, 8 + 2 * len(screen_factors)).value = f"={cY}{r}-{cFIT}{r}"
        wr.cell(r, 8 + 2 * len(screen_factors)).number_format = PCT
        wr.cell(r, 9 + 2 * len(screen_factors)).value = f"={cU}{r}^2"
        wr.cell(r, 9 + 2 * len(screen_factors)).number_format = SCI

    # worked-name specific block
    sr = fac_row + len(design) + 3
    wr.cell(sr, 1, "Worked name — specific (idiosyncratic) risk").font = bold
    wr.cell(sr + 1, 1, f"{fig2tkr.get(worked,'')} residual u (day d)")
    wr.cell(sr + 1, 2, f"={cU}{base + wi}").number_format = PCT
    wr.cell(sr + 2, 1, "u²  (one day's contribution to SpecificVar)")
    wr.cell(sr + 2, 2, f"={cU2}{base + wi}").number_format = SCI
    wr.cell(sr + 3, 1, "SpecificVar in frame (EWMA of u² across the month, ref)")
    sv_ref = sv_frame[(sv_frame["Position"] == worked)]
    sv_ref = sv_ref[sv_ref["Date"] == (M + pd.offsets.MonthEnd(0))]
    if len(sv_ref):
        wr.cell(sr + 3, 2, float(sv_ref["SpecificVar"].iloc[0])).number_format = SCI
    wr.cell(sr + 4, 1, "(u² is one day's input to that EWMA, not the smoothed value)")
    wr.column_dimensions["A"].width = 18

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)

    # tie-out report (pandas mirror vs frame) so we know the formulas target the right numbers
    print(f"wrote {XLSX}")
    print(f"  month-end M = {M.date()}   regression day d = {d1.date()}")
    print(f"  worked name = {fig2tkr.get(worked)} ({worked})   cross-section: {nrows} names raw, "
          f"{len(figs)} in regression, {len(screen_factors)} factors survive screen")
    print("  §3b tie-out (worked-name loading: pandas mirror vs production frame):")
    for f in STYLE_FACTORS:
        mref = loadM.loc[worked, f] if worked in loadM.index else np.nan
        fv = fr_load.get(f)
        if pd.notna(mref) and fv is not None:
            print(f"    {f:11} mirror {mref:+.4f}   frame {fv:+.4f}   Δ {abs(mref-fv):.1e}")
    print("  §3e tie-out (factor return on d: pandas LINEST mirror vs frame):")
    for f, bcoef in zip(fac_names, beta):
        fv = fr_ref.get(f)
        if fv is not None:
            print(f"    {f:11} mirror {bcoef:+.5f}   frame {fv:+.5f}   Δ {abs(bcoef-fv):.1e}")


if __name__ == "__main__":
    build()
